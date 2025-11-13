"""
Report Manager Module
Gönderilen maillerin Excel raporunu yönetir.
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ReportManager:
    """Mail raporu yönetim sınıfı"""
    
    def __init__(self, report_file: str = "data/mail_raporu.xlsx"):
        self.report_file = Path(report_file)
        self.df = None
        
        # Rapor dosyasını yükle veya oluştur
        self._initialize_report()
    
    def _initialize_report(self):
        """Rapor dosyasını başlat"""
        try:
            if self.report_file.exists():
                # Mevcut raporu yükle
                self.df = pd.read_excel(self.report_file)
                logger.info(f"✅ Mevcut rapor yüklendi: {len(self.df)} kayıt")
            else:
                # Yeni rapor oluştur
                self.df = pd.DataFrame(columns=[
                    'Gönderim Tarihi',
                    'Gönderim Saati',
                    'İhale No',
                    'İhale Adı',
                    'Yönetici',
                    'Yönetici Mail',
                    'Hatırlatma Tipi',
                    'Kalan Gün',
                    'Başlangıç Tarihi',
                    'Durum',
                    'Hata Mesajı',
                    'Retry Sayısı'
                ])
                
                # Klasörü oluştur
                self.report_file.parent.mkdir(parents=True, exist_ok=True)
                
                # İlk kez kaydet
                self.df.to_excel(self.report_file, index=False)
                logger.info(f"✅ Yeni rapor dosyası oluşturuldu: {self.report_file}")
                
        except Exception as e:
            logger.error(f"❌ Rapor başlatma hatası: {str(e)}")
            raise
    
    def add_entries(self, email_results: list, reminder_info: dict = None) -> dict:
        """
        Birden fazla mail sonucunu rapora ekle
        
        Args:
            email_results: Email gönderim sonuçları
            reminder_info: Ek hatırlatma bilgileri (opsiyonel)
            
        Returns:
            dict: Ekleme sonucu
        """
        try:
            added_count = 0
            errors = []
            
            for result in email_results:
                # Entry oluştur
                entry = {
                    'Gönderim Tarihi': result["timestamp"].strftime("%Y-%m-%d"),
                    'Gönderim Saati': result["timestamp"].strftime("%H:%M:%S"),
                    'İhale No': result["ihale_no"],
                    'İhale Adı': result["ihale_adi"],
                    'Yönetici': self._extract_manager_name(result["recipient"]),
                    'Yönetici Mail': result["recipient"],
                    'Hatırlatma Tipi': reminder_info.get("hatirlatma_tipi", "") if reminder_info else "",
                    'Kalan Gün': reminder_info.get("kalan_gun", 0) if reminder_info else 0,
                    'Başlangıç Tarihi': reminder_info.get("baslangic_tarihi", "") if reminder_info else "",
                    'Durum': "Başarılı" if result["status"] == "sent" else "Başarısız",
                    'Hata Mesajı': result.get("error_message", ""),
                    'Retry Sayısı': result.get("retry_count", 0)
                }
                
                # DataFrame'e ekle
                self.df = pd.concat([self.df, pd.DataFrame([entry])], ignore_index=True)
                added_count += 1
            
            # Dosyaya kaydet
            self._save_report()
            
            logger.info(f"✅ {added_count} kayıt rapora eklendi")
            
            return {
                "success": True,
                "entries_added": added_count,
                "total_entries": len(self.df),
                "errors": errors
            }
            
        except Exception as e:
            logger.error(f"❌ Rapor ekleme hatası: {str(e)}")
            return {
                "success": False,
                "entries_added": 0,
                "total_entries": len(self.df),
                "errors": [str(e)]
            }
    
    def add_entry(self, email_result: dict, reminder: dict) -> bool:
        """
        Tek bir mail sonucunu rapora ekle
        
        Args:
            email_result: Email gönderim sonucu
            reminder: Hatırlatma bilgileri
            
        Returns:
            bool: Başarı durumu
        """
        try:
            # Entry oluştur
            entry = {
                'Gönderim Tarihi': email_result["timestamp"].strftime("%Y-%m-%d"),
                'Gönderim Saati': email_result["timestamp"].strftime("%H:%M:%S"),
                'İhale No': email_result["ihale_no"],
                'İhale Adı': email_result["ihale_adi"],
                'Yönetici': reminder["yonetici"],
                'Yönetici Mail': email_result["recipient"],
                'Hatırlatma Tipi': reminder["hatirlatma_tipi"],
                'Kalan Gün': reminder["kalan_gun"],
                'Başlangıç Tarihi': reminder["baslangic_tarihi"].strftime("%Y-%m-%d"),
                'Durum': "Başarılı" if email_result["status"] == "sent" else "Başarısız",
                'Hata Mesajı': email_result.get("error_message", ""),
                'Retry Sayısı': email_result.get("retry_count", 0)
            }
            
            # DataFrame'e ekle
            self.df = pd.concat([self.df, pd.DataFrame([entry])], ignore_index=True)
            
            # Dosyaya kaydet
            self._save_report()
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Entry ekleme hatası: {str(e)}")
            return False
    
    def _save_report(self):
        """Raporu Excel dosyasına kaydet"""
        try:
            # Excel writer ile formatting yap
            with pd.ExcelWriter(self.report_file, engine='openpyxl') as writer:
                self.df.to_excel(writer, index=False, sheet_name='Mail Raporu')
                
                # Worksheet'i al
                worksheet = writer.sheets['Mail Raporu']
                
                # Column widths ayarla
                column_widths = {
                    'A': 15,  # Gönderim Tarihi
                    'B': 12,  # Gönderim Saati
                    'C': 10,  # İhale No
                    'D': 40,  # İhale Adı
                    'E': 20,  # Yönetici
                    'F': 30,  # Yönetici Mail
                    'G': 15,  # Hatırlatma Tipi
                    'H': 12,  # Kalan Gün
                    'I': 15,  # Başlangıç Tarihi
                    'J': 12,  # Durum
                    'K': 40,  # Hata Mesajı
                    'L': 12   # Retry Sayısı
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Header'ı bold yap
                from openpyxl.styles import Font, PatternFill
                
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                
                # Başarılı/Başarısız durumları renklendir
                for row in range(2, len(self.df) + 2):
                    durum_cell = worksheet[f'J{row}']
                    if durum_cell.value == "Başarılı":
                        durum_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif durum_cell.value == "Başarısız":
                        durum_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            logger.info(f"✅ Rapor kaydedildi: {self.report_file}")
            
        except Exception as e:
            logger.error(f"❌ Rapor kaydetme hatası: {str(e)}")
            raise
    
    def _extract_manager_name(self, email: str) -> str:
        """Email adresinden yönetici ismini çıkar (varsa)"""
        # Bu basit bir implementasyon
        # Gerçek uygulamada daha sofistike bir yöntem kullanılabilir
        return ""
    
    def get_daily_statistics(self, date: datetime = None) -> dict:
        """
        Belirli bir günün istatistiklerini al
        
        Args:
            date: Tarih (None ise bugün)
            
        Returns:
            dict: Günlük istatistikler
        """
        try:
            if date is None:
                date = datetime.now()
            
            date_str = date.strftime("%Y-%m-%d")
            
            # Günlük filtreleme
            daily_df = self.df[self.df['Gönderim Tarihi'] == date_str]
            
            if len(daily_df) == 0:
                return {
                    "tarih": date_str,
                    "toplam_gonderim": 0,
                    "basarili": 0,
                    "basarisiz": 0,
                    "60_gun": 0,
                    "30_gun": 0,
                    "1_gun": 0,
                    "benzersiz_yonetici": 0
                }
            
            return {
                "tarih": date_str,
                "toplam_gonderim": len(daily_df),
                "basarili": len(daily_df[daily_df['Durum'] == 'Başarılı']),
                "basarisiz": len(daily_df[daily_df['Durum'] == 'Başarısız']),
                "60_gun": len(daily_df[daily_df['Hatırlatma Tipi'] == '60_gun']),
                "30_gun": len(daily_df[daily_df['Hatırlatma Tipi'] == '30_gun']),
                "1_gun": len(daily_df[daily_df['Hatırlatma Tipi'] == '1_gun']),
                "benzersiz_yonetici": daily_df['Yönetici Mail'].nunique()
            }
            
        except Exception as e:
            logger.error(f"❌ İstatistik hesaplama hatası: {str(e)}")
            return {}
    
    def get_failed_reports(self, limit: int = 10) -> list:
        """
        Başarısız gönderileri listele
        
        Args:
            limit: Maksimum sonuç sayısı
            
        Returns:
            list: Başarısız kayıtlar
        """
        try:
            failed_df = self.df[self.df['Durum'] == 'Başarısız'].tail(limit)
            return failed_df.to_dict('records')
            
        except Exception as e:
            logger.error(f"❌ Başarısız rapor listeleme hatası: {str(e)}")
            return []
    
    def backup_report(self) -> bool:
        """Rapor dosyasının yedeğini al"""
        try:
            backup_dir = Path("data/backups")
            backup_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = backup_dir / f"mail_raporu_backup_{timestamp}.xlsx"
            
            if self.report_file.exists():
                import shutil
                shutil.copy2(self.report_file, backup_path)
                logger.info(f"✅ Rapor backup oluşturuldu: {backup_path}")
                return True
            
            return False
            
        except Exception as e:
            logger.error(f"❌ Rapor backup hatası: {str(e)}")
            return False


if __name__ == "__main__":
    # Test
    report_manager = ReportManager()
    
    # Test entry
    test_result = {
        "ihale_no": 9,
        "ihale_adi": "Test İhalesi",
        "recipient": "test@example.com",
        "status": "sent",
        "timestamp": datetime.now(),
        "error_message": None,
        "retry_count": 0
    }
    
    test_reminder = {
        "yonetici": "Test Yönetici",
        "hatirlatma_tipi": "30_gun",
        "kalan_gun": 30,
        "baslangic_tarihi": datetime.now()
    }
    
    # Test ekle
    result = report_manager.add_entry(test_result, test_reminder)
    print(f"\n{'✅' if result else '❌'} Test entry eklendi")
    
    # İstatistikleri göster
    stats = report_manager.get_daily_statistics()
    print(f"\n📊 Bugünkü İstatistikler:")
    print(f"  Toplam: {stats.get('toplam_gonderim', 0)}")
    print(f"  Başarılı: {stats.get('basarili', 0)}")
    print(f"  Başarısız: {stats.get('basarisiz', 0)}")
